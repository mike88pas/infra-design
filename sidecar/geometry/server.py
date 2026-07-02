#!/usr/bin/env python3
"""
Infra Design — sidecar geometrii (ezdxf / Shapely / A*).

Protokół: newline-delimited JSON przez stdio.
    request : {"id": int, "method": str, "params": dict}\n   (stdin)
    response: {"id": int, "ok": true, "result": any}\n         (stdout)
            | {"id": int, "ok": false, "error": str}\n          (stdout)

stdout jest zarezerwowany WYŁĄCZNIE dla protokołu (jeden JSON na linię).
Diagnostyka idzie na stderr.

Metody:
    ping            — handshake (wersja ezdxf/Pythona)              [F0]
    import_dxf      — wczytanie rzutu DXF → warstwy + encje + bbox  [F1]
    polygonize      — wykrycie pomieszczeń z segmentów ścian        [F1]
    extract_devices — INSERT-y (symbole urządzeń) → warstwa+poz+atr [F2]
"""

import json
import os
import sys
import platform

# Katalog tego pliku na sys.path — by `import safepath` działał zarówno przy
# uruchomieniu jako skrypt (`python server.py`), jak i przy ładowaniu modułu
# przez importlib w testach (spec_from_file_location nie dodaje katalogu).
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))


def _ezdxf_version() -> str:
    try:
        import ezdxf  # type: ignore
        return getattr(ezdxf, "__version__", "unknown")
    except Exception as exc:  # noqa: BLE001
        return f"missing ({exc.__class__.__name__})"


HANDLERS = {}


def handler(name):
    def deco(fn):
        HANDLERS[name] = fn
        return fn
    return deco


# ── cache parsowania DXF ─────────────────────────────────────────────────────
#
# Plik 40 MB parsuje się ~10 s. Kreator woła kilka handlerów na tym samym pliku
# w jednym imporcie — bez cache każdy płaci pełny koszt. Trzymamy DOKŁADNIE JEDEN
# zparsowany dokument w pamięci (maszyna ma napięty RAM — nigdy 2 doce naraz).
# Klucz: (abspath, mtime_ns, size) — invaliduje się sam, gdy plik się zmieni.
# Wszystkie handlery są READ-ONLY, więc współdzielenie doc między wywołaniami jest bezpieczne.

_DOC_CACHE_KEY = None
_DOC_CACHE = None


def _load_doc(path):
    """Zwraca zcache'owany ezdxf `Drawing` dla `path` albo parsuje od nowa.

    `path` powinno być już zwalidowane (safepath.validate_in_path). Cache keyujemy
    po (os.path.abspath, mtime_ns, size). Przed wczytaniem nowego doca zwalniamy
    poprzedni (None), by nie trzymać dwóch naraz.
    """
    import ezdxf  # type: ignore

    global _DOC_CACHE_KEY, _DOC_CACHE
    abspath = os.path.abspath(path)
    st = os.stat(abspath)
    key = (abspath, st.st_mtime_ns, st.st_size)
    if key == _DOC_CACHE_KEY and _DOC_CACHE is not None:
        return _DOC_CACHE
    # Zwolnij poprzedni dokument PRZED wczytaniem nowego (napięty RAM — bez 2 doców naraz).
    _DOC_CACHE = None
    _DOC_CACHE_KEY = None
    doc = ezdxf.readfile(path)
    _DOC_CACHE = doc
    _DOC_CACHE_KEY = key
    return doc


# ── Pomocnicze ──────────────────────────────────────────────────────────────

# Ochrona pamięci: twardy limit encji renderowalnych z jednego DXF.
MAX_ENTITIES = 300_000

# Górny limit segmentów dla polygonize. Powyżej tej liczby Shapely (unary_union +
# polygonize) na maszynie z napiętym RAM potrafi paść — bail z czytelnym błędem.
POLYGONIZE_MAX_SEGMENTS = 20_000

# $INSUNITS → jednostka modelu. Schema TS zna tylko mm|m; cm/in mapujemy na mm.
_INSUNITS_M = {6: "m"}  # 6 = metry; reszta traktowana jak mm


def _aci_to_hex(aci, default="#c8c8c8") -> str:
    """Kolor ACI (indeks AutoCAD) → hex. BYLAYER/BYBLOCK → domyślny."""
    try:
        from ezdxf import colors as ezcolors  # type: ignore
        if aci is None or aci in (0, 256):
            return default
        r, g, b = ezcolors.aci2rgb(abs(int(aci)))
        return f"#{r:02x}{g:02x}{b:02x}"
    except Exception:  # noqa: BLE001
        return default


class _BBox:
    """Akumulator prostokąta otaczającego."""

    __slots__ = ("minx", "miny", "maxx", "maxy", "_set")

    def __init__(self):
        self.minx = self.miny = float("inf")
        self.maxx = self.maxy = float("-inf")
        self._set = False

    def add(self, x: float, y: float) -> None:
        if x < self.minx:
            self.minx = x
        if y < self.miny:
            self.miny = y
        if x > self.maxx:
            self.maxx = x
        if y > self.maxy:
            self.maxy = y
        self._set = True

    def to_json(self) -> dict:
        if not self._set:
            return {"minX": 0.0, "minY": 0.0, "maxX": 0.0, "maxY": 0.0}
        return {
            "minX": self.minx,
            "minY": self.miny,
            "maxX": self.maxx,
            "maxY": self.maxy,
        }

    def diagonal(self) -> float:
        if not self._set:
            return 0.0
        dx = self.maxx - self.minx
        dy = self.maxy - self.miny
        return (dx * dx + dy * dy) ** 0.5


# ── ping ────────────────────────────────────────────────────────────────────

@handler("ping")
def _ping(_params):
    return {
        "pong": True,
        "ezdxf": _ezdxf_version(),
        "python": platform.python_version(),
    }


# ── import_dxf ──────────────────────────────────────────────────────────────

@handler("import_dxf")
def _import_dxf(params):
    """
    params: { "path": str }
    return: DxfDocument { layers, entities, bbox, units, entityCount, truncated? }
    """
    import ezdxf  # type: ignore
    import safepath

    path = safepath.validate_in_path(params.get("path"), params.get("_allowedRoots"))
    # Limit encji do RENDEROWANIA. Podkłady zwektoryzowane z PDF mają dziesiątki tysięcy
    # mikro-encji — pełny zrzut zatyka IPC i renderer. Caller (kreator) może podać niski
    # limit (podkład jest tylko poglądowy; projekt liczy się z wykazu pomieszczeń).
    max_render = int(params.get("maxRenderEntities", MAX_ENTITIES) or MAX_ENTITIES)
    max_render = max(1, min(max_render, MAX_ENTITIES))

    doc = _load_doc(path)
    msp = doc.modelspace()

    units = _INSUNITS_M.get(int(doc.header.get("$INSUNITS", 0) or 0), "mm")

    # Warstwy (kolor: true-color jeśli jest, inaczej z ACI)
    layers = []
    for lay in doc.layers:
        rgb = lay.rgb if lay.rgb else None
        if rgb:
            color = "#{:02x}{:02x}{:02x}".format(*rgb)
        else:
            color = _aci_to_hex(lay.color)
        layers.append({
            "name": lay.dxf.name,
            "color": color,
            "visible": not bool(lay.is_off()),
        })

    bbox = _BBox()
    entities = []
    truncated = 0

    def cap() -> bool:
        nonlocal truncated
        if len(entities) >= max_render:
            truncated += 1
            return True
        return False

    for e in msp:
        dxftype = e.dxftype()
        layer = getattr(e.dxf, "layer", "0")
        try:
            if dxftype == "LINE":
                if cap():
                    continue
                a, b = e.dxf.start, e.dxf.end
                bbox.add(a.x, a.y)
                bbox.add(b.x, b.y)
                entities.append({
                    "t": "line", "layer": layer,
                    "a": {"x": a.x, "y": a.y},
                    "b": {"x": b.x, "y": b.y},
                })
            elif dxftype == "LWPOLYLINE":
                if cap():
                    continue
                pts = [{"x": x, "y": y} for x, y in e.get_points("xy")]
                for p in pts:
                    bbox.add(p["x"], p["y"])
                entities.append({
                    "t": "polyline", "layer": layer,
                    "pts": pts, "closed": bool(e.closed),
                })
            elif dxftype == "POLYLINE":
                if cap():
                    continue
                pts = [{"x": v.dxf.location.x, "y": v.dxf.location.y} for v in e.vertices]
                for p in pts:
                    bbox.add(p["x"], p["y"])
                entities.append({
                    "t": "polyline", "layer": layer,
                    "pts": pts, "closed": bool(e.is_closed),
                })
            elif dxftype == "CIRCLE":
                if cap():
                    continue
                c, r = e.dxf.center, float(e.dxf.radius)
                bbox.add(c.x - r, c.y - r)
                bbox.add(c.x + r, c.y + r)
                entities.append({
                    "t": "circle", "layer": layer,
                    "c": {"x": c.x, "y": c.y}, "r": r,
                })
            elif dxftype == "ARC":
                if cap():
                    continue
                c, r = e.dxf.center, float(e.dxf.radius)
                bbox.add(c.x - r, c.y - r)
                bbox.add(c.x + r, c.y + r)
                entities.append({
                    "t": "arc", "layer": layer,
                    "c": {"x": c.x, "y": c.y}, "r": r,
                    "start": float(e.dxf.start_angle),
                    "end": float(e.dxf.end_angle),
                })
            elif dxftype == "INSERT":
                if cap():
                    continue
                p = e.dxf.insert
                bbox.add(p.x, p.y)
                entities.append({
                    "t": "insert", "layer": layer,
                    "at": {"x": p.x, "y": p.y},
                    "name": e.dxf.name,
                    "rotation": float(e.dxf.rotation or 0.0),
                    "sx": float(e.dxf.xscale or 1.0),
                    "sy": float(e.dxf.yscale or 1.0),
                })
            elif dxftype in ("TEXT", "MTEXT"):
                if cap():
                    continue
                if dxftype == "TEXT":
                    p = e.dxf.insert
                    txt = e.dxf.text
                    h = float(e.dxf.height or 0.0)
                else:
                    p = e.dxf.insert
                    txt = e.plain_text()
                    h = float(e.dxf.char_height or 0.0)
                bbox.add(p.x, p.y)
                entities.append({
                    "t": "text", "layer": layer,
                    "at": {"x": p.x, "y": p.y},
                    "text": txt, "height": h,
                })
        except Exception as exc:  # noqa: BLE001 — pojedyncza wadliwa encja nie wywala importu
            print(f"[import_dxf] pomijam {dxftype}: {exc}", file=sys.stderr, flush=True)

    result = {
        "layers": layers,
        "entities": entities,
        "bbox": bbox.to_json(),
        "units": units,
        "entityCount": len(entities),
    }
    if truncated:
        result["truncated"] = truncated
    return result


# ── pomocnicze: ściany w blokach ─────────────────────────────────────────────

# Maks. głębokość rekurencji przy eksplozji bloków (ochrona przed zagnieżdżeniem/cyklem).
_MAX_EXPLODE_DEPTH = 6


def _layer_matches(layer: str, wall_set) -> bool:
    """Czy warstwa należy do zbioru ścian (None = wszystkie).

    Dopasowanie po podłańcuchu: encje eksplodowane z bloku mają warstwy typu
    '<blok>$0$A-WALL', więc token 'A-WALL' (czy 'WALL') z `wallLayers` je złapie.
    """
    if wall_set is None:
        return True
    if layer in wall_set:
        return True
    return any(w in layer for w in wall_set)


def _iter_exploded(entity, explode, depth=0):
    """Yield-uje encje-liście: gdy `explode` i to INSERT — rekurencyjnie wnętrze
    bloku (w WCS przez virtual_entities); w przeciwnym razie samą encję. Podkład
    architektoniczny bywa wstawiony jako jeden blok — bez eksplozji jego wnętrza nie widać.
    """
    if entity.dxftype() == "INSERT" and explode and depth < _MAX_EXPLODE_DEPTH:
        try:
            for sub in entity.virtual_entities():
                yield from _iter_exploded(sub, explode, depth + 1)
        except Exception as exc:  # noqa: BLE001 — wadliwy blok nie wywala importu
            print(f"[explode] pominięty blok: {exc}", file=sys.stderr, flush=True)
    else:
        yield entity


def _iter_wall_geometry(entity, explode, depth=0):
    """Jak `_iter_exploded`, ale tylko geometria ścian (LINE/LWPOLYLINE/POLYLINE)."""
    for g in _iter_exploded(entity, explode, depth):
        if g.dxftype() in ("LINE", "LWPOLYLINE", "POLYLINE"):
            yield g


_MTEXT_UNICODE_RE = None


def _decode_mtext(text: str) -> str:
    """Dekoduje escapy '\\U+XXXX' (polskie znaki w MTEXT, np. ł=U+0142) i czyści białe znaki.

    Standard AutoCAD = 4 cyfry hex ('\\U+0142'). Niektóre konwertery (DWG zwektoryzowany
    z PDF) gubią wiodące zera i zapisują 3 hex ('\\U+142'). Próbujemy 4-hex, potem 3-hex.
    """
    import re
    global _MTEXT_UNICODE_RE
    if _MTEXT_UNICODE_RE is None:
        _MTEXT_UNICODE_RE = re.compile(r"\\U\+([0-9A-Fa-f]{4})|\\U\+([0-9A-Fa-f]{3})")
    return _MTEXT_UNICODE_RE.sub(
        lambda m: chr(int(m.group(1) or m.group(2), 16)), text
    ).strip()


# ── polygonize ──────────────────────────────────────────────────────────────

@handler("polygonize")
def _polygonize(params):
    """
    Buduje zamknięte pomieszczenia z segmentów (LINE/LWPOLYLINE) wskazanych warstw.
    Obsługuje „brudny" DXF (niedomknięte narożniki) snapowaniem do siatki.

    params: {
        "path": str,                 # ścieżka DXF
        "wallLayers": [str]?,        # warstwy ścian (dopasowanie po podłańcuchu; domyślnie: wszystkie)
        "explodeBlocks": bool?,      # wejdź w bloki INSERT (ściany w podkładzie); domyślnie False
        "snap": float?,              # tolerancja snapu (jedn. modelu); auto z bbox
        "minArea": float?            # min. pole pomieszczenia (jedn.^2)
    }
    return: PolygonizeResult { polygons:[{points,area}], snapTolerance }
    """
    import ezdxf  # type: ignore
    from shapely.geometry import LineString  # type: ignore
    from shapely.ops import unary_union, polygonize  # type: ignore
    import safepath

    path = safepath.validate_in_path(params.get("path"), params.get("_allowedRoots"))

    wall_layers = params.get("wallLayers")
    wall_set = set(wall_layers) if wall_layers else None
    explode = bool(params.get("explodeBlocks", False))

    doc = _load_doc(path)
    msp = doc.modelspace()

    # Zbierz segmenty + oszacuj rozmiar rysunku (do auto-snapu)
    segments = []  # list[((x1,y1),(x2,y2))]
    bbox = _BBox()

    def emit(geom) -> None:
        t = geom.dxftype()
        if t == "LINE":
            a, b = geom.dxf.start, geom.dxf.end
            segments.append(((a.x, a.y), (b.x, b.y)))
            bbox.add(a.x, a.y)
            bbox.add(b.x, b.y)
        else:  # LWPOLYLINE / POLYLINE
            if t == "LWPOLYLINE":
                pts = [(x, y) for x, y in geom.get_points("xy")]
                closed = bool(geom.closed)
            else:
                pts = [(v.dxf.location.x, v.dxf.location.y) for v in geom.vertices]
                closed = bool(geom.is_closed)
            if closed and len(pts) > 2:
                pts = pts + [pts[0]]
            for i in range(len(pts) - 1):
                segments.append((pts[i], pts[i + 1]))
            for x, y in pts:
                bbox.add(x, y)

    for e in msp:
        if len(segments) > POLYGONIZE_MAX_SEGMENTS:
            break
        for geom in _iter_wall_geometry(e, explode):
            if _layer_matches(getattr(geom.dxf, "layer", "0"), wall_set):
                emit(geom)

    # Ochrona: polygonize na dziesiątkach tysięcy segmentów (np. podkład zwektoryzowany
    # z PDF, bez warstwy ścian) potrafi zżreć pamięć i ubić sidecar. Bezpieczny bail z
    # czytelną wskazówką, zamiast OOM/zawisu — szczególnie na maszynach z napiętym RAM.
    if len(segments) > POLYGONIZE_MAX_SEGMENTS:
        return {
            "polygons": [],
            "snapTolerance": 0.0,
            "error": (
                f"Za zlozona geometria ({len(segments)}+ segmentow) — to wyglada na podklad "
                "wektoryzowany z PDF lub brak wskazanej warstwy scian. Zawez 'warstwy scian' "
                "albo uzyj trybu 'Tabela Zestawienie' zamiast 'ze scian'."
            ),
        }

    if not segments:
        return {"polygons": [], "snapTolerance": 0.0}

    diag = bbox.diagonal()
    snap = params.get("snap")
    if snap is None:
        # ~0.1% przekątnej rysunku — domyka typowe luki narożników brudnego DXF
        snap = max(diag * 0.001, 1e-9)
    snap = float(snap)

    def snapped(pt):
        if snap <= 0:
            return pt
        return (round(pt[0] / snap) * snap, round(pt[1] / snap) * snap)

    lines = []
    for a, b in segments:
        sa, sb = snapped(a), snapped(b)
        if sa != sb:
            lines.append(LineString([sa, sb]))

    if not lines:
        return {"polygons": [], "snapTolerance": snap}

    merged = unary_union(lines)
    faces = list(polygonize(merged))

    min_area = params.get("minArea")
    if min_area is None:
        # odfiltruj artefakty: < ~0.01% pola bboxa
        area_bbox = max((bbox.maxx - bbox.minx) * (bbox.maxy - bbox.miny), 1.0)
        min_area = area_bbox * 1e-4
    min_area = float(min_area)

    polygons = []
    for poly in faces:
        if poly.area < min_area:
            continue
        coords = list(poly.exterior.coords)[:-1]  # bez powtórzonego pkt zamykającego
        polygons.append({
            "points": [{"x": x, "y": y} for x, y in coords],
            "area": float(poly.area),
        })

    polygons.sort(key=lambda p: p["area"], reverse=True)
    return {"polygons": polygons, "snapTolerance": snap}


# ── extract_devices ──────────────────────────────────────────────────────────

@handler("extract_devices")
def _extract_devices(params):
    """
    Zwraca symbole urządzeń (bloki INSERT z modelspace) wraz z warstwą, pozycją,
    obrotem i atrybutami (ATTRIB, np. IDFX/NR — przypisanie portu do szafy).

    Mapowanie warstwa→system/typ robi strona TS (src/domain/dxf/systemMapping.ts) —
    symbole bywają blokami anonimowymi (*U34), więc klasyfikacja idzie po WARSTWIE.

    params: {
        "path": str,
        "layers": [str]?,        # filtr po podłańcuchu nazwy warstwy (domyślnie: wszystkie)
        "includeAttribs": bool?  # dołącz atrybuty ATTRIB (domyślnie True)
    }
    return: { inserts: [{layer,name,at,rotation,sx,sy,attribs}], count }
    """
    import ezdxf  # type: ignore

    import safepath
    path = safepath.validate_in_path(params.get("path"), params.get("_allowedRoots"))

    layer_filter = params.get("layers")
    include_attribs = params.get("includeAttribs", True)

    doc = _load_doc(path)
    msp = doc.modelspace()

    inserts = []
    for ins in msp.query("INSERT"):
        layer = getattr(ins.dxf, "layer", "0")
        if layer_filter and not any(f in layer for f in layer_filter):
            continue
        p = ins.dxf.insert
        attribs = {}
        if include_attribs:
            try:
                for a in ins.attribs:
                    attribs[a.dxf.tag] = a.dxf.text
            except Exception:  # noqa: BLE001 — brak/wadliwe atrybuty nie blokują encji
                pass
        inserts.append({
            "layer": layer,
            "name": ins.dxf.name,
            "at": {"x": p.x, "y": p.y},
            "rotation": float(ins.dxf.rotation or 0.0),
            "sx": float(ins.dxf.xscale or 1.0),
            "sy": float(ins.dxf.yscale or 1.0),
            "attribs": attribs,
        })

    return {"inserts": inserts, "count": len(inserts)}


# ── extract_rooms ────────────────────────────────────────────────────────────

# Numer pomieszczenia: 1.11, 1.13A, 0.2 ; pole: "224.64 m²", "3,09 m2"
_ROOM_NUM_RE = None
_ROOM_AREA_RE = None


def _parse_room_label(parts):
    """Z listy linii MTEXT (numer/nazwa/pole) buduje (number, name, areaM2)."""
    import re
    global _ROOM_NUM_RE, _ROOM_AREA_RE
    if _ROOM_NUM_RE is None:
        _ROOM_NUM_RE = re.compile(r"^\d+\.\w+$")
        _ROOM_AREA_RE = re.compile(r"([\d ]+[.,]\d+)\s*m[²2]?", re.I)
    number, area, name_parts = "", None, []
    for p in parts:
        p = p.strip()
        if not p:
            continue
        am = _ROOM_AREA_RE.search(p)
        if am and area is None:
            area = float(am.group(1).replace(" ", "").replace(",", "."))
            continue
        if not number and _ROOM_NUM_RE.match(p):
            number = p
            continue
        name_parts.append(p)
    return number, " ".join(name_parts), area


@handler("extract_rooms")
def _extract_rooms(params):
    """
    Wykaz pomieszczeń z etykiet pól (warstwy A-AREA itp.): numer + nazwa + oficjalny
    metraż (m²) nadane przez architekta. To czystsze źródło niż rekonstrukcja ze ścian.

    Ramka etykiety to zamknięta polilinia; numer/nazwa/pole to MTEXT-y w jej obrysie.

    params: {
        "path": str,
        "areaLayers": [str]?,     # podłańcuchy warstw pól (domyślnie ['AREA'])
        "explodeBlocks": bool?    # wejdź w bloki (domyślnie True — etykiety bywają w podkładzie)
    }
    return: { rooms: [{number, name, areaM2, at:{x,y}, tag:[{x,y}]}], count }
    """
    import ezdxf  # type: ignore
    from shapely.geometry import Polygon, Point as SPoint  # type: ignore

    import safepath
    path = safepath.validate_in_path(params.get("path"), params.get("_allowedRoots"))

    area_layers = params.get("areaLayers") or ["AREA"]
    explode = params.get("explodeBlocks", True)

    doc = _load_doc(path)
    msp = doc.modelspace()

    def on_area(layer: str) -> bool:
        return any(a in layer for a in area_layers)

    tags = []  # list[(Polygon, [pts])]
    texts = []  # list[(x, y, str)]
    for e in msp:
        for g in _iter_exploded(e, explode):
            layer = getattr(g.dxf, "layer", "0")
            if not on_area(layer):
                continue
            t = g.dxftype()
            if t in ("LWPOLYLINE", "POLYLINE"):
                try:
                    if t == "LWPOLYLINE":
                        pts = [(x, y) for x, y in g.get_points("xy")]
                        closed = bool(g.closed)
                    else:
                        pts = [(v.dxf.location.x, v.dxf.location.y) for v in g.vertices]
                        closed = bool(g.is_closed)
                    if closed and len(pts) >= 3:
                        tags.append((Polygon(pts), pts))
                except Exception:  # noqa: BLE001
                    pass
            elif t in ("TEXT", "MTEXT"):
                try:
                    raw = g.plain_text() if t == "MTEXT" else g.dxf.text
                    p = g.dxf.insert
                    texts.append((p.x, p.y, _decode_mtext(raw)))
                except Exception:  # noqa: BLE001
                    pass

    rooms = []
    seen = set()
    for poly, pts in tags:
        inside = [txt for (x, y, txt) in texts if poly.contains(SPoint(x, y))]
        if not inside:
            continue
        # Spłaszcz wieloliniowe MTEXT-y (numer/nazwa/pole bywają w jednym lub osobnych).
        lines = [ln for txt in inside for ln in txt.split("\n")]
        number, name, area_m2 = _parse_room_label(lines)
        if not name and not number:
            continue
        key = number or name
        if key in seen:
            continue
        seen.add(key)
        c = poly.centroid
        rooms.append({
            "number": number,
            "name": name,
            "areaM2": area_m2,
            "at": {"x": c.x, "y": c.y},
            "tag": [{"x": x, "y": y} for x, y in pts],
        })

    rooms.sort(key=lambda r: (r["areaM2"] or 0), reverse=True)
    return {"rooms": rooms, "count": len(rooms)}


# ── extract_rooms_schedule ────────────────────────────────────────────────────
#
# Dla rzutów bez warstw pól (np. DWG zwektoryzowany z PDF): wykaz pomieszczeń jest
# w TABELI "Zestawienie" (kolumny: numer | nazwa | powierzchnia m²), a na rzucie są
# tylko ETYKIETY-NUMERY (0.14, 1.07…) na środku pomieszczeń. Łączymy oba po numerze:
#   tabela  → number → (name, areaM2)
#   rzut    → number → pozycja (at)
# Separacja: numery pomieszczeń mają KROPKĘ (0.14), powierzchnie PRZECINEK (65,99).

_SCHED_NUM_RE = None
_SCHED_AREA_RE = None


def _collect_texts(msp, explode=True):
    """Zbiera (text, x, y) z TEXT i MTEXT (z rozbiciem bloków)."""
    out = []
    for e in msp:
        for g in _iter_exploded(e, explode):
            t = g.dxftype()
            if t not in ("TEXT", "MTEXT"):
                continue
            try:
                raw = g.plain_text() if t == "MTEXT" else g.dxf.text
                txt = _decode_mtext(raw).strip()
                if not txt:
                    continue
                p = g.dxf.insert
                out.append((txt, float(p.x), float(p.y)))
            except Exception:  # noqa: BLE001
                pass
    return out


@handler("extract_rooms_schedule")
def _extract_rooms_schedule(params):
    """
    Wykaz pomieszczeń z TABELI zestawienia + etykiet-numerów na rzucie.

    params: {
        "path": str,
        "explodeBlocks": bool?,   # domyślnie True
        "scale": float?,          # mnożnik pozycji (np. 0.1 dla 1:100 mm→m); domyślnie 1.0
        "headerName": str?,       # tekst nagłówka kolumny nazwy (domyślnie 'Pomieszczenie')
        "headerArea": str?        # tekst nagłówka kolumny pola (domyślnie 'Powierzchnia')
    }
    return: { rooms:[{number,name,areaM2,at,tag}], count, table_rows, plan_labels, unmatched }
    """
    import re
    import ezdxf  # type: ignore

    global _SCHED_NUM_RE, _SCHED_AREA_RE
    if _SCHED_NUM_RE is None:
        _SCHED_NUM_RE = re.compile(r"^\d+\.\w{1,3}$")          # 0.14, 1.07, 3.2A
        _SCHED_AREA_RE = re.compile(r"^([\d ]+,\d+)\s*(?:m[²2]?)?$")  # 65,99 / 1 203,11

    import safepath
    path = safepath.validate_in_path(params.get("path"), params.get("_allowedRoots"))
    explode = params.get("explodeBlocks", True)
    scale = float(params.get("scale", 1.0))
    hdr_name = params.get("headerName", "Pomieszczenie")
    hdr_area = params.get("headerArea", "Powierzchnia")

    doc = _load_doc(path)
    texts = _collect_texts(doc.modelspace(), explode)

    # 1) Nagłówki tabeli → x kolumn + y nagłówka.
    name_x = area_x = header_y = None
    for txt, x, y in texts:
        if txt == hdr_name:
            name_x, header_y = x, y
        elif txt == hdr_area:
            area_x = x
    if name_x is None or area_x is None:
        return {"rooms": [], "count": 0, "error": "Nie znaleziono nagłówków tabeli zestawienia"}

    # Granica tabela / rzut: numery tabeli są tuż przy/na lewo od kolumny nazwy.
    table_left = name_x - 60.0   # numery rzutu mają x < table_left

    # 2) Wiersze tabeli (y <= header_y): grupuj po Y (tolerancja 2.5).
    in_table = [(t, x, y) for (t, x, y) in texts if y <= header_y + 1.0 and x >= table_left]
    in_table.sort(key=lambda r: -r[2])
    rows = []  # list[dict(y, number, name, area)]
    for t, x, y in in_table:
        row = None
        for r in rows:
            if abs(r["y"] - y) <= 2.5:
                row = r
                break
        if row is None:
            row = {"y": y, "number": "", "name": "", "area": None}
            rows.append(row)
        if _SCHED_NUM_RE.match(t) and x < name_x - 8:
            row["number"] = t
        elif _SCHED_AREA_RE.match(t):
            am = _SCHED_AREA_RE.match(t)
            try:
                row["area"] = float(am.group(1).replace(" ", "").replace(",", "."))
            except ValueError:
                pass
        elif x < area_x - 15 and not _SCHED_NUM_RE.match(t):
            row["name"] = (row["name"] + " " + t).strip()

    sched = {}  # number → (name, area)
    for r in rows:
        if r["number"]:
            sched[r["number"]] = (r["name"], r["area"])

    # 3) Etykiety-numery na rzucie (x < table_left) → pozycja (uśredniona przy duplikatach).
    plan = {}
    for t, x, y in texts:
        if x < table_left and _SCHED_NUM_RE.match(t):
            plan.setdefault(t, []).append((x, y))
    plan_pos = {n: (sum(p[0] for p in pts) / len(pts), sum(p[1] for p in pts) / len(pts))
                for n, pts in plan.items()}

    # 4) Join po numerze. Pozycje skalowane (scale). Brak pozycji → siatka zastępcza.
    rooms = []
    unmatched = []
    keys = sorted(set(sched) | set(plan_pos), key=lambda k: (len(k), k))
    gx = gy = 0.0
    for i, num in enumerate(keys):
        name, area = sched.get(num, ("", None))
        if num in plan_pos:
            px, py = plan_pos[num]
            at = {"x": px * scale, "y": py * scale}
        else:
            at = {"x": (gx := gx + 5.0), "y": gy}   # zastępcza siatka (brak etykiety na rzucie)
            unmatched.append(num)
        s = 0.5
        rooms.append({
            "number": num,
            "name": name,
            "areaM2": area,
            "at": at,
            "tag": [
                {"x": at["x"] - s, "y": at["y"] - s},
                {"x": at["x"] + s, "y": at["y"] - s},
                {"x": at["x"] + s, "y": at["y"] + s},
                {"x": at["x"] - s, "y": at["y"] + s},
            ],
        })
    rooms.sort(key=lambda r: (r["areaM2"] or 0), reverse=True)
    return {
        "rooms": rooms,
        "count": len(rooms),
        "table_rows": len(sched),
        "plan_labels": len(plan_pos),
        "unmatched": unmatched,
    }


# ── route_cables (A*) ─────────────────────────────────────────────────────────

# Siatka A*: górny limit komórek na dłuższym boku (ochrona pamięci/czasu).
_ROUTE_MAX_CELLS_SIDE = 220


def _dijkstra_multi(seeds, blocked, w, h, ortho=True):
    """Multi-source Dijkstra od celów (szaf) po całej wolnej siatce — JEDNO przejście.
    seeds: list[(targetIndex, (gx,gy))]. Zwraca (dist, parent, origin) per komórka.
    Każde urządzenie odczytuje potem swój koszt i trasę backtrace — bez N osobnych A*.

    `ortho=True` (domyślnie): 4-sąsiedztwo — kable pod kątem prostym, jak realne koryta/trasy
    kablowe (czysty rysunek + policzalne koryta). `ortho=False`: 8-sąsiedztwo (skróty po skosie).
    """
    import heapq

    dist = {}
    parent = {}
    origin = {}
    heap = []
    for idx, c in seeds:
        if c in blocked:
            continue
        if c not in dist:
            dist[c] = 0.0
            parent[c] = None
            origin[c] = idx
            heapq.heappush(heap, (0.0, c))
    if ortho:
        nbrs = ((1, 0), (-1, 0), (0, 1), (0, -1))
    else:
        nbrs = ((1, 0), (-1, 0), (0, 1), (0, -1), (1, 1), (1, -1), (-1, 1), (-1, -1))
    while heap:
        d, cur = heapq.heappop(heap)
        if d > dist.get(cur, float("inf")):
            continue
        cx, cy = cur
        for dx, dy in nbrs:
            nx, ny = cx + dx, cy + dy
            if nx < 0 or ny < 0 or nx >= w or ny >= h:
                continue
            nb = (nx, ny)
            if nb in blocked:
                continue
            nd = d + (1.41421356 if (dx and dy) else 1.0)
            if nd < dist.get(nb, float("inf")):
                dist[nb] = nd
                parent[nb] = cur
                origin[nb] = origin[cur]
                heapq.heappush(heap, (nd, nb))
    return dist, parent, origin


def _simplify_cells(cells):
    """Usuwa środkowe komórki współliniowych biegów → polilinia tylko z punktami załamań."""
    if len(cells) <= 2:
        return cells
    out = [cells[0]]
    for i in range(1, len(cells) - 1):
        ax, ay = cells[i - 1]
        bx, by = cells[i]
        cx, cy = cells[i + 1]
        if (bx - ax, by - ay) == (cx - bx, cy - by):
            continue  # ten sam kierunek — punkt pośredni pomijamy
        out.append(cells[i])
    out.append(cells[-1])
    return out


@handler("route_cables")
def _route_cables(params):
    """
    Trasuje kable od źródeł (urządzeń) do najbliższego celu (szafy) algorytmem A*
    po siatce z przeszkodami ze ścian. Brak przejścia → linia prosta (fallback).
    Długości w JEDNOSTKACH MODELU (TS przelicza na metry przez unitMm).

    params: {
        "path": str,
        "sources": [{x,y}],         # urządzenia
        "targets": [{x,y}],         # szafy/rozdzielnie
        "wallLayers": [str]?,       # ściany-przeszkody (dopasowanie po podłańcuchu)
        "explodeBlocks": bool?,     # wejdź w bloki (domyślnie True)
        "cell": float?,             # rozmiar komórki (auto z bboxa)
        "inflate": int?             # pogrubienie ścian w komórkach (domyślnie 0)
    }
    return: { routes: [{sourceIndex, targetIndex, path:[{x,y}], length, method}], cell, grid:{w,h} }
    """
    import ezdxf  # type: ignore

    import safepath
    sources = params.get("sources") or []
    targets = params.get("targets") or []
    if not sources or not targets:
        return {"routes": [], "cell": 0.0, "grid": {"w": 0, "h": 0}}
    path = safepath.validate_in_path(params.get("path"), params.get("_allowedRoots"))

    wall_layers = params.get("wallLayers")
    wall_set = set(wall_layers) if wall_layers else None
    door_layers = params.get("doorLayers")
    door_set = set(door_layers) if door_layers else None
    door_clear = int(params.get("doorClear", 2))
    explode = params.get("explodeBlocks", True)
    inflate = int(params.get("inflate", 0))

    doc = _load_doc(path)
    msp = doc.modelspace()

    # Segmenty ścian + bbox całości (źródła, cele, ściany)
    segments = []
    bbox = _BBox()
    for s in sources + targets:
        bbox.add(s["x"], s["y"])
    for e in msp:
        for geom in _iter_wall_geometry(e, explode):
            if not _layer_matches(getattr(geom.dxf, "layer", "0"), wall_set):
                continue
            t = geom.dxftype()
            if t == "LINE":
                a, b = geom.dxf.start, geom.dxf.end
                segments.append(((a.x, a.y), (b.x, b.y)))
                bbox.add(a.x, a.y)
                bbox.add(b.x, b.y)
            else:
                if t == "LWPOLYLINE":
                    pts = [(x, y) for x, y in geom.get_points("xy")]
                    closed = bool(geom.closed)
                else:
                    pts = [(v.dxf.location.x, v.dxf.location.y) for v in geom.vertices]
                    closed = bool(geom.is_closed)
                if closed and len(pts) > 2:
                    pts = pts + [pts[0]]
                for i in range(len(pts) - 1):
                    segments.append((pts[i], pts[i + 1]))

    # Otwory drzwiowe: geometria drzwi (do "przebicia" w ścianach po rasteryzacji),
    # żeby kabel przeszedł przez drzwi, a nie skrótem przez mur.
    door_segments = []
    if door_set is not None:
        for e in msp:
            for geom in _iter_wall_geometry(e, explode):
                if not _layer_matches(getattr(geom.dxf, "layer", "0"), door_set):
                    continue
                t = geom.dxftype()
                if t == "LINE":
                    a, b = geom.dxf.start, geom.dxf.end
                    door_segments.append(((a.x, a.y), (b.x, b.y)))
                else:
                    if t == "LWPOLYLINE":
                        pts = [(x, y) for x, y in geom.get_points("xy")]
                        closed = bool(geom.closed)
                    else:
                        pts = [(v.dxf.location.x, v.dxf.location.y) for v in geom.vertices]
                        closed = bool(geom.is_closed)
                    if closed and len(pts) > 2:
                        pts = pts + [pts[0]]
                    for i in range(len(pts) - 1):
                        door_segments.append((pts[i], pts[i + 1]))

    minx, miny = bbox.minx, bbox.miny
    width = max(bbox.maxx - minx, 1.0)
    height = max(bbox.maxy - miny, 1.0)

    cell = params.get("cell")
    if cell is None:
        cell = max(width, height) / _ROUTE_MAX_CELLS_SIDE
    cell = float(max(cell, 1e-6))

    w = int(width / cell) + 2
    h = int(height / cell) + 2

    def to_cell(p):
        return (int((p["x"] - minx) / cell), int((p["y"] - miny) / cell))

    def to_cell_xy(x, y):
        return (int((x - minx) / cell), int((y - miny) / cell))

    # Rasteryzacja ścian → blocked
    blocked = set()
    for (ax, ay), (bx, by) in segments:
        steps = int((abs(bx - ax) + abs(by - ay)) / cell) + 1
        for i in range(steps + 1):
            t = i / steps
            blocked.add(to_cell_xy(ax + (bx - ax) * t, ay + (by - ay) * t))
    if inflate > 0:
        extra = set()
        for (cx, cy) in blocked:
            for dx in range(-inflate, inflate + 1):
                for dy in range(-inflate, inflate + 1):
                    extra.add((cx + dx, cy + dy))
        blocked = extra

    # Przebij otwory drzwiowe: usuń komórki drzwi (z marginesem door_clear) z przeszkód.
    # Robione PO inflate, by pogrubienie ścian nie zasklepiło z powrotem przejścia.
    if door_segments:
        clear = set()
        for (ax, ay), (bx, by) in door_segments:
            steps = int((abs(bx - ax) + abs(by - ay)) / cell) + 1
            for i in range(steps + 1):
                tt = i / steps
                dcx, dcy = to_cell_xy(ax + (bx - ax) * tt, ay + (by - ay) * tt)
                for dx in range(-door_clear, door_clear + 1):
                    for dy in range(-door_clear, door_clear + 1):
                        clear.add((dcx + dx, dcy + dy))
        blocked -= clear

    # Jedno przejście Dijkstry od wszystkich szaf po wolnej siatce.
    seeds = []
    for ti, t in enumerate(targets):
        tc = to_cell(t)
        seeds.append((ti, tc if tc not in blocked else _nearest_free(tc, blocked, w, h)))
    seeds = [(ti, c) for ti, c in seeds if c is not None]
    ortho = params.get("ortho", True)
    dist, parent, origin = _dijkstra_multi(seeds, blocked, w, h, ortho=ortho)

    def to_point(c):
        return {"x": minx + (c[0] + 0.5) * cell, "y": miny + (c[1] + 0.5) * cell}

    routes = []
    for si, s in enumerate(sources):
        sc = to_cell(s)
        start = sc if (sc in dist) else _nearest_free(sc, blocked, w, h)
        if start is not None and start in dist:
            cells = [start]
            cur = start
            while parent.get(cur) is not None:
                cur = parent[cur]
                cells.append(cur)
            length = dist[start] * cell
            cells = _simplify_cells(cells)  # tylko punkty załamań → czysta polilinia
            routes.append({
                "sourceIndex": si, "targetIndex": origin[start],
                "path": [to_point(c) for c in cells], "length": length, "method": "astar",
            })
        else:
            # brak przejścia (źródło odcięte) → linia prosta do najbliższego celu
            ti = min(range(len(targets)), key=lambda j: (targets[j]["x"] - s["x"]) ** 2 + (targets[j]["y"] - s["y"]) ** 2)
            tgt = targets[ti]
            length = ((tgt["x"] - s["x"]) ** 2 + (tgt["y"] - s["y"]) ** 2) ** 0.5
            routes.append({
                "sourceIndex": si, "targetIndex": ti,
                "path": [{"x": s["x"], "y": s["y"]}, {"x": tgt["x"], "y": tgt["y"]}],
                "length": length, "method": "straight",
            })

    return {"routes": routes, "cell": cell, "grid": {"w": w, "h": h}}


def _nearest_free(cell, blocked, w, h, max_r=8):
    """Najbliższa wolna komórka wokół `cell` (gdy źródło/cel wpadło na ścianę)."""
    cx, cy = cell
    if 0 <= cx < w and 0 <= cy < h and cell not in blocked:
        return cell
    for r in range(1, max_r + 1):
        for dx in range(-r, r + 1):
            for dy in range(-r, r + 1):
                nx, ny = cx + dx, cy + dy
                if 0 <= nx < w and 0 <= ny < h and (nx, ny) not in blocked:
                    return (nx, ny)
    return None


# ── export_dxf ────────────────────────────────────────────────────────────

# Warstwy rysunku instalacji (nazwa → kolor ACI).
_EXPORT_LAYERS = {
    "INSTAL-LAN": 5, "INSTAL-CCTV": 1, "INSTAL-KD": 3, "INSTAL-AP": 4,
    "INSTAL-TRASY": 8, "INSTAL-KORYTA": 251, "INSTAL-SZAFY": 2, "INSTAL-OPIS": 7,
    "INSTAL-LEGENDA": 7,
}


def _device_layer(system: str, type_key: str) -> str:
    if type_key.startswith("lan.ap"):
        return "INSTAL-AP"
    return {"lan": "INSTAL-LAN", "cctv": "INSTAL-CCTV", "kd": "INSTAL-KD"}.get(system, "INSTAL-OPIS")


def _draw_symbol(msp, system, type_key, x, y, s):
    """Symbol urządzenia: AP=okrąg, CCTV=trójkąt, reszta=kwadrat (bok 2·s)."""
    layer = _device_layer(system, type_key)
    if type_key.startswith("lan.ap"):
        msp.add_circle((x, y), s, dxfattribs={"layer": layer})
    elif system == "cctv":
        msp.add_lwpolyline([(x - s, y - s), (x + s, y - s), (x, y + s)], close=True, dxfattribs={"layer": layer})
    else:
        msp.add_lwpolyline(
            [(x - s, y - s), (x + s, y - s), (x + s, y + s), (x - s, y + s)],
            close=True, dxfattribs={"layer": layer},
        )


@handler("export_dxf")
def _export_dxf(params):
    """
    Zapisuje rysunek instalacji do DXF (overlay nakładany na podkład w CAD; docelowo XREF).
    Symbole per system, trasy jako polilinie, etykiety pomieszczeń, szafy, legenda + tabelka.

    params: {
        "path": str,                         # plik wyjściowy .dxf
        "devices": [{system,typeKey,position:{x,y}}],
        "routes": [{path:[{x,y}], system}],
        "trays": [{path:[{x,y}], widthDraw, widthMm}]?,  # koryta (magistrale)
        "rooms": [{name, at:{x,y}}],
        "cabinets": [{x,y}],
        "legend": [{label, count}],
        "meta": {project, drawing, designer, license},
        "symbolSize": float?                 # połowa boku symbolu (domyślnie 250 mm)
    }
    return: { path, devices, routes, trays }
    """
    import ezdxf  # type: ignore
    import safepath

    out = safepath.validate_out_path(params.get("path"), params.get("_allowedRoots"))
    devices = params.get("devices", [])
    routes = params.get("routes", [])
    trays = params.get("trays", []) or []
    rooms = params.get("rooms", [])
    cabinets = params.get("cabinets", [])
    legend = params.get("legend", [])
    meta = params.get("meta", {})
    s = float(params.get("symbolSize", 250))

    doc = ezdxf.new("R2018", setup=True)
    doc.header["$INSUNITS"] = 4  # mm
    msp = doc.modelspace()
    for name, color in _EXPORT_LAYERS.items():
        if name not in doc.layers:
            doc.layers.add(name, color=color)

    bbox = _BBox()

    # Koryta kablowe (najgłębiej — magistrale pod trasami). LWPOLYLINE z natywną
    # szerokością (const_width) + etykieta "KORYTO {mm}" wzdłuż najdłuższego segmentu.
    import math as _math
    for t in trays:
        pts = [(q["x"], q["y"]) for q in t.get("path", [])]
        if len(pts) < 2:
            continue
        attribs = {"layer": "INSTAL-KORYTA"}
        w = float(t.get("widthDraw", 0) or 0)
        if w > 0:
            attribs["const_width"] = w
        msp.add_lwpolyline(pts, dxfattribs=attribs)
        # najdłuższy segment → środek + kąt etykiety
        bi, bl = 0, -1.0
        for i in range(len(pts) - 1):
            L = _math.hypot(pts[i + 1][0] - pts[i][0], pts[i + 1][1] - pts[i][1])
            if L > bl:
                bl, bi = L, i
        mx = (pts[bi][0] + pts[bi + 1][0]) / 2
        my = (pts[bi][1] + pts[bi + 1][1]) / 2
        ang = _math.degrees(_math.atan2(pts[bi + 1][1] - pts[bi][1], pts[bi + 1][0] - pts[bi][0]))
        if ang > 90 or ang < -90:
            ang += 180  # tekst zawsze "do góry nogami" → obróć
        msp.add_text(
            f"KORYTO {int(t.get('widthMm', 0))}", height=s * 0.9,
            dxfattribs={"layer": "INSTAL-KORYTA", "rotation": ang},
        ).set_placement((mx, my))
        for x, y in pts:
            bbox.add(x, y)

    # Trasy (pod spodem)
    for r in routes:
        pts = [(q["x"], q["y"]) for q in r.get("path", [])]
        if len(pts) >= 2:
            msp.add_lwpolyline(pts, dxfattribs={"layer": "INSTAL-TRASY"})
            for x, y in pts:
                bbox.add(x, y)

    # Urządzenia
    for d in devices:
        p = d["position"]
        _draw_symbol(msp, d.get("system", ""), d.get("typeKey", ""), p["x"], p["y"], s)
        bbox.add(p["x"], p["y"])

    # Szafy/IDF — większy kwadrat z opisem
    for c in cabinets:
        x, y = c["x"], c["y"]
        msp.add_lwpolyline(
            [(x - s * 2, y - s * 2), (x + s * 2, y - s * 2), (x + s * 2, y + s * 2), (x - s * 2, y + s * 2)],
            close=True, dxfattribs={"layer": "INSTAL-SZAFY"},
        )
        msp.add_text("IDF", height=s * 1.2, dxfattribs={"layer": "INSTAL-SZAFY"}).set_placement((x - s * 1.5, y))
        bbox.add(x, y)

    # Etykiety pomieszczeń
    for rm in rooms:
        at = rm.get("at", {})
        if "x" in at:
            msp.add_text(rm.get("name", ""), height=s * 0.9, dxfattribs={"layer": "INSTAL-OPIS"}).set_placement(
                (at["x"], at["y"])
            )

    # Legenda + tabelka projektu — na prawo od rysunku
    if bbox._set:
        lx = bbox.maxx + s * 12
        ly = bbox.maxy
        lh = s * 5  # odstęp wierszy
        msp.add_text("LEGENDA", height=s * 2.4, dxfattribs={"layer": "INSTAL-LEGENDA"}).set_placement((lx, ly))
        row = ly - lh * 1.5
        for item in legend:
            txt = f"{item.get('label', '')} — {item.get('count', 0)} szt"
            msp.add_text(txt, height=s * 1.8, dxfattribs={"layer": "INSTAL-LEGENDA"}).set_placement((lx + s * 4, row))
            row -= lh
        # Tabelka projektu
        row -= lh
        for line in [
            f"Projekt: {meta.get('project', '')}",
            f"Rysunek: {meta.get('drawing', '')}",
            f"Projektant: {meta.get('designer', '')}  upr. {meta.get('license', '')}",
            "Podpis: ........................................",
            "UWAGA: rysunek wspomaga projektanta — nie zastępuje autoryzacji projektu.",
        ]:
            msp.add_text(line, height=s * 1.6, dxfattribs={"layer": "INSTAL-LEGENDA"}).set_placement((lx, row))
            row -= lh

    doc.saveas(out)
    return {"path": out, "devices": len(devices), "routes": len(routes), "trays": len(trays)}


@handler("export_kosztorys")
def _export_kosztorys(params):
    """
    Zapisuje kosztorys/zestawienie inwestorskie do XLSX w formacie klienta.

    params: {
        "path": str,                          # plik wyjściowy .xlsx
        "kosztorys": {                        # struktura z buildKosztorys (TS)
            "categories": [{label, kosztorys:[...], zestawienie:[...], netto, brutto}],
            "all": [{lp,sku,name,unit,qty,price,netto,brutto}],
            "total": {netto, brutto},
            "vatPct": int,
            "meta": {project, generatedNote}
        }
    }
    Arkusze: KOSZTORYS CAŁOŚĆ + ZESTAWIENIE CAŁOŚĆ, potem para Kosztorys/Zestawienie per kategoria.
    return: { path, sheets, rows }
    """
    import openpyxl  # type: ignore
    from openpyxl.styles import Font  # type: ignore
    import safepath

    out = safepath.validate_out_path(params.get("path"), params.get("_allowedRoots"), allowed_ext=(".xlsx",))
    k = params.get("kosztorys", {}) or {}
    categories = k.get("categories", [])
    all_rows = k.get("all", [])
    total = k.get("total", {}) or {}
    meta = k.get("meta", {}) or {}

    KH = ["Lp.", "Towar", "Ilość", "Cena", "Waluta", "Netto", "Brutto", "Nazwa"]
    ZH = ["Lp.", "Towar", "Ilość", "J.M", "Nazwa"]
    bold = Font(bold=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def kosztorys_sheet(title, rows):
        ws = wb.create_sheet(title[:31])
        ws.append(["Pozycje oferty"])
        ws["A1"].font = bold
        ws.append(KH)
        for c in ws[2]:
            c.font = bold
        for r in rows:
            ws.append([
                r.get("lp"), r.get("sku"), r.get("qty"), r.get("price"), "PLN",
                r.get("netto"), r.get("brutto"), r.get("name"),
            ])
        return ws

    def zestawienie_sheet(title, rows):
        ws = wb.create_sheet(title[:31])
        ws.append(ZH)
        for c in ws[1]:
            c.font = bold
        for r in rows:
            ws.append([r.get("lp"), r.get("sku"), r.get("qty"), r.get("unit"), r.get("name")])
        return ws

    # CAŁOŚĆ — wszystkie pozycje (Lp ciągłe)
    wsc = kosztorys_sheet("KOSZTORYS CAŁOŚĆ", all_rows)
    wsc.append([])
    wsc.append(["", "RAZEM", "", "", "PLN", total.get("netto"), total.get("brutto"), "Suma"])
    for c in wsc[wsc.max_row]:
        c.font = bold
    z_all = [
        {"lp": i + 1, "sku": r.get("sku"), "qty": r.get("qty"), "unit": r.get("unit"), "name": r.get("name")}
        for i, r in enumerate(all_rows)
    ]
    zestawienie_sheet("ZESTAWIENIE CAŁOŚĆ", z_all)

    # Para arkuszy per kategoria
    for cat in categories:
        label = cat.get("label", "?")
        kosztorys_sheet(f"Kosztorys {label}", cat.get("kosztorys", []))
        zestawienie_sheet(f"Zestawienie {label}", cat.get("zestawienie", []))

    # auto-szerokości (zgrubnie)
    for ws in wb.worksheets:
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 8), 60)

    wb.save(out)
    return {"path": out, "sheets": len(wb.worksheets), "rows": len(all_rows)}


@handler("export_rack_elevation")
def _export_rack_elevation(params):
    """
    Rysuje elewację (widok od frontu) szaf 19" do DXF — na wzór „Widok/Elewacja szaf".

    params: {
        "path": str,                           # plik wyjściowy .dxf
        "racks": [{name, uHeight, units:[{uPos,uSize,label}]}],
        "meta": {project, designer, license}
    }
    return: { path, racks, units }
    """
    import ezdxf  # type: ignore
    import safepath

    out = safepath.validate_out_path(params.get("path"), params.get("_allowedRoots"))
    racks = params.get("racks", [])
    meta = params.get("meta", {})

    U = 44.45            # wysokość 1U [mm]
    INNER = 482.6        # szerokość montażowa 19" [mm]
    FRAME_W = 600.0      # szerokość obrysu szafy
    GAP = 400.0          # odstęp między szafami
    MARGIN = (FRAME_W - INNER) / 2

    doc = ezdxf.new("R2018", setup=True)
    doc.header["$INSUNITS"] = 4  # mm
    msp = doc.modelspace()
    for name, color in (("RACK-FRAME", 7), ("RACK-UNIT", 5), ("RACK-TEXT", 3), ("RACK-EMPTY", 8)):
        if name not in doc.layers:
            doc.layers.add(name, color=color)

    def rect(x0, y0, x1, y1, layer):
        msp.add_lwpolyline([(x0, y0), (x1, y0), (x1, y1), (x0, y1)], close=True, dxfattribs={"layer": layer})

    total_units = 0
    x = 0.0
    for rack in racks:
        uH = int(rack.get("uHeight", 42))
        H = uH * U
        # Obrys szafy + szyny montażowe
        rect(x, 0, x + FRAME_W, H, "RACK-FRAME")
        rect(x + MARGIN, 0, x + MARGIN + INNER, H, "RACK-FRAME")
        # Nazwa szafy
        msp.add_text(rack.get("name", "Szafa"), height=U * 0.6,
                     dxfattribs={"layer": "RACK-TEXT"}).set_placement((x, H + U * 0.6))
        # Numeracja U (co 1U, od dołu) + pozioma kreska siatki
        for u in range(uH):
            y = u * U
            msp.add_text(str(uH - u), height=U * 0.35,
                         dxfattribs={"layer": "RACK-TEXT"}).set_placement((x - U * 0.9, y + U * 0.3))
        # Zajęte U
        occupied = set()
        for unit in rack.get("units", []):
            up = int(unit.get("uPos", 1))
            us = int(unit.get("uSize", 1))
            # rysujemy od dołu: uPos=1 na dole
            y0 = (up - 1) * U
            y1 = y0 + us * U
            rect(x + MARGIN, y0, x + MARGIN + INNER, y1, "RACK-UNIT")
            msp.add_text(unit.get("label", ""), height=U * 0.4,
                         dxfattribs={"layer": "RACK-TEXT"}).set_placement((x + MARGIN + U * 0.4, y0 + us * U * 0.3))
            for k in range(us):
                occupied.add(up + k)
            total_units += 1
        # Puste U (szrafowanie cienką linią)
        for u in range(1, uH + 1):
            if u not in occupied:
                y0 = (u - 1) * U
                rect(x + MARGIN, y0, x + MARGIN + INNER, y0 + U, "RACK-EMPTY")
        x += FRAME_W + GAP

    # Tabelka projektu
    if racks:
        ty = -U * 2
        for line in [
            f"Projekt: {meta.get('project', '')}",
            f"Projektant: {meta.get('designer', '')}  upr. {meta.get('license', '')}",
            "Elewacja szaf — rysunek wspomaga projektanta, nie zastępuje autoryzacji projektu.",
        ]:
            msp.add_text(line, height=U * 0.45, dxfattribs={"layer": "RACK-TEXT"}).set_placement((0, ty))
            ty -= U

    doc.saveas(out)
    return {"path": out, "racks": len(racks), "units": total_units}


# ── pętla protokołu ─────────────────────────────────────────────────────────

def dispatch(method: str, params: dict):
    fn = HANDLERS.get(method)
    if fn is None:
        raise ValueError(f"Nieznana metoda: {method}")
    return fn(params or {})


def _result_size(method: str, result) -> str:
    """Krótki opis rozmiaru wyniku do logu czasowego (debug)."""
    if not isinstance(result, dict):
        return ""
    for k in ("entities", "routes", "rooms", "polygons", "inserts"):
        v = result.get(k)
        if isinstance(v, list):
            extra = f" truncated={result['truncated']}" if result.get("truncated") else ""
            return f" {k}={len(v)}{extra}"
    return ""


def main() -> int:
    import time
    # Protokół JSON jest UTF-8 — na Windows domyślne kodowanie stdio to cp125x, co
    # rozsypuje polskie znaki w danych z requestu (np. nazwy „MODUŁ"). Wymuszamy UTF-8.
    for stream in (sys.stdin, sys.stdout):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[attr-defined]
        except Exception:  # noqa: BLE001
            pass
    debug = bool(os.environ.get("INFRA_DEBUG"))
    print("infra-design sidecar gotowy", file=sys.stderr, flush=True)
    for line in sys.stdin:
        line = line.strip()
        if not line:
            continue
        method = "?"
        t0 = time.perf_counter()
        try:
            msg = json.loads(line)
            msg_id = msg.get("id")
            method = msg.get("method", "")
            if debug:
                print(f"[timing] -> {method} start", file=sys.stderr, flush=True)
            result = dispatch(method, msg.get("params"))
            out = {"id": msg_id, "ok": True, "result": result}
            if debug:
                dt = (time.perf_counter() - t0) * 1000
                print(f"[timing] <- {method} OK {dt:.0f}ms{_result_size(method, result)}",
                      file=sys.stderr, flush=True)
        except Exception as exc:  # noqa: BLE001
            if debug:
                dt = (time.perf_counter() - t0) * 1000
                print(f"[timing] <- {method} ERROR {dt:.0f}ms: {exc.__class__.__name__}: {exc}",
                      file=sys.stderr, flush=True)
            out = {
                "id": locals().get("msg_id"),
                "ok": False,
                "error": f"{exc.__class__.__name__}: {exc}",
            }
        sys.stdout.write(json.dumps(out) + "\n")
        sys.stdout.flush()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
